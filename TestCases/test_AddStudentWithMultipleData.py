import requests
import pytest
import json
import jsonpath
import openpyxl


base_url = "http://thetestingworldapi.com/"


def test_add_multiple_students():
    # API access
    api_url = base_url + "api/studentsDetails"
    file = open("C:\\Users\\SridharRaju\\Desktop\\API_tasks\\add_newStudent.json.txt", "r")
    json_request = json.loads(file.read())

    # Excel actions
    workbook = openpyxl.load_workbook("C:\\Users\\SridharRaju\\Desktop\\API_tasks\\newStudentDetail.xlsx")
    sheet = workbook['Sheet1']
    rows = sheet.max_row
    for i in range(1, rows+1):
        first_name_cell = sheet.cell(row=i, column=1)
        middle_name_cell = sheet.cell(row=i, column=1)
        last_name_cell = sheet.cell(row=i, column=1)
        date_of_birth_cell = sheet.cell(row=i, column=1)
        json_request['first_name'] = first_name_cell.value
        json_request['middle_name'] = middle_name_cell.value
        json_request['last_name'] = last_name_cell.value
        json_request['date_of_birth'] = date_of_birth_cell.value
        response = requests.post(api_url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201